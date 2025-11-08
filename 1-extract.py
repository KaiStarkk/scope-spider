import json
import sys
from pathlib import Path
from openpyxl import load_workbook

HEADER_ROW_INDEX = 3
TICKER_COL_INDEX = 0
COMPANY_COL_INDEX = 1


def main():
    if len(sys.argv) < 2:
        sys.exit("Usage: 1-extract.py <xlsx_file> [>output_file.json]")

    xlsx = Path(sys.argv[1])
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb.active
    if ws is None:
        sys.exit("Error: Could not access worksheet")
    rows = list(ws.iter_rows(values_only=True))
    data = [
        {
            "company_name": str(row[COMPANY_COL_INDEX]).strip(),
            "ticker": str(row[TICKER_COL_INDEX]).strip(),
        }
        for row in rows[HEADER_ROW_INDEX + 1 :]
        if row[TICKER_COL_INDEX] and row[COMPANY_COL_INDEX]
    ]
    wb.close()
    print(json.dumps(data, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
