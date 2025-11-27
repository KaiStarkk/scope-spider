from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

ROOT_DIR = Path(__file__).resolve().parent.parent
DEFAULT_COMPANIES_PATH = ROOT_DIR / "companies.json"
DEFAULT_WORKBOOK_PATH = ROOT_DIR / "ASX ESG Screening.xlsx"

HEADER_ROW = 4
DATA_START_ROW = HEADER_ROW + 1


@dataclass(frozen=True)
class DerivedMetrics:
    profitability_ratio: Optional[float]
    reputational_ratio: Optional[float]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Append derived metrics to the Companies_Results sheet."
    )
    parser.add_argument(
        "--companies",
        type=Path,
        default=DEFAULT_COMPANIES_PATH,
        help="Path to companies.json (default: %(default)s).",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=DEFAULT_WORKBOOK_PATH,
        help="Path to the Excel workbook that should be updated (default: %(default)s).",
    )
    parser.add_argument(
        "--sheet",
        default="Companies_Results",
        help="Worksheet name to update (default: %(default)s).",
    )
    parser.add_argument(
        "--profitability-header",
        default="Profitability Ratio",
        help="Column header label for net income / revenue (default: %(default)s).",
    )
    parser.add_argument(
        "--reputation-header",
        default="Reputational Concern Ratio",
        help="Column header label for net zero mentions / revenue (default: %(default)s).",
    )
    return parser.parse_args()


def _normalise_name(value: Optional[str]) -> Optional[str]:
    if not value:
        return None
    text = value.strip()
    return text.lower() if text else None


def load_metrics(companies_path: Path) -> Tuple[Dict[str, DerivedMetrics], Dict[str, DerivedMetrics]]:
    payload = json.loads(companies_path.read_text(encoding="utf-8"))
    companies = payload.get("companies")
    if not isinstance(companies, list):
        raise ValueError("companies.json payload must contain a 'companies' list.")

    by_ticker: Dict[str, DerivedMetrics] = {}
    by_name: Dict[str, DerivedMetrics] = {}

    for company in companies:
        identity = company.get("identity") or {}
        annotations = company.get("annotations") or {}
        ticker = identity.get("ticker")
        name = identity.get("name")
        profitability = annotations.get("profitability_ratio")
        reputation = annotations.get("reputational_concern_ratio")
        if profitability is None and reputation is None:
            continue
        metrics = DerivedMetrics(
            profitability_ratio=float(profitability) if profitability is not None else None,
            reputational_ratio=float(reputation) if reputation is not None else None,
        )
        if isinstance(ticker, str):
            token = ticker.strip().upper()
            if token:
                by_ticker[token] = metrics
        norm_name = _normalise_name(name if isinstance(name, str) else None)
        if norm_name:
            by_name[norm_name] = metrics

    return by_ticker, by_name


def update_sheet(
    workbook_path: Path,
    sheet_name: str,
    headers: Tuple[str, str],
    by_ticker: Dict[str, DerivedMetrics],
    by_name: Dict[str, DerivedMetrics],
) -> int:
    workbook = load_workbook(workbook_path)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Worksheet '{sheet_name}' not found in {workbook_path}.")
    sheet = workbook[sheet_name]

    start_column = sheet.max_column + 1
    profitability_col = start_column
    reputational_col = start_column + 1

    # Assign headers on the main header row; keep multi-row structure by leaving earlier rows blank.
    sheet.cell(row=HEADER_ROW, column=profitability_col, value=headers[0])
    sheet.cell(row=HEADER_ROW, column=reputational_col, value=headers[1])

    updates = 0
    for row_idx in range(DATA_START_ROW, sheet.max_row + 1):
        ticker = sheet.cell(row=row_idx, column=1).value
        name = sheet.cell(row=row_idx, column=2).value
        metrics = None
        if isinstance(ticker, str):
            metrics = by_ticker.get(ticker.strip().upper())
        if metrics is None and isinstance(name, str):
            metrics = by_name.get(_normalise_name(name))
        if metrics is None:
            continue
        sheet.cell(row=row_idx, column=profitability_col, value=metrics.profitability_ratio)
        sheet.cell(row=row_idx, column=reputational_col, value=metrics.reputational_ratio)
        updates += 1

    workbook.save(workbook_path)
    return updates, (get_column_letter(profitability_col), get_column_letter(reputational_col))


def main() -> int:
    args = parse_args()
    if not args.workbook.exists():
        raise SystemExit(f"[error] Workbook not found: {args.workbook}")
    if not args.companies.exists():
        raise SystemExit(f"[error] companies.json not found: {args.companies}")

    by_ticker, by_name = load_metrics(args.companies)
    profitability_label = args.profitability_header.strip() or "Profitability Ratio"
    reputation_label = args.reputation_header.strip() or "Reputational Concern Ratio"
    updates, column_letters = update_sheet(
        workbook_path=args.workbook,
        sheet_name=args.sheet,
        headers=(profitability_label, reputation_label),
        by_ticker=by_ticker,
        by_name=by_name,
    )
    print(
        f"[info] Updated {updates} rows in '{args.sheet}' with derived metrics. "
        f"Columns added at {column_letters[0]} and {column_letters[1]}."
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
