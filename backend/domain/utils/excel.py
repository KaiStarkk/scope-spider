from __future__ import annotations

from pathlib import Path
from typing import List

from openpyxl import load_workbook

from ..models import Company, Identity


def extract_companies_from_workbook(
    xlsx_path: Path,
    *,
    header_row_index: int,
    ticker_col_index: int,
    company_col_index: int,
) -> List[Company]:
    workbook = load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        sheet = workbook.active
        if sheet is None:
            raise ValueError("Error: Could not access worksheet")
        rows = list(sheet.iter_rows(values_only=True))
        companies: List[Company] = []
        for row in rows[header_row_index + 1 :]:
            ticker_value = row[ticker_col_index]
            company_value = row[company_col_index]
            if (
                ticker_value is None
                or company_value is None
                or not str(ticker_value).strip()
                or not str(company_value).strip()
            ):
                continue
            companies.append(
                Company(
                    identity=Identity(
                        name=str(company_value).strip(),
                        ticker=str(ticker_value).strip(),
                    ),
                )
            )
        return companies
    finally:
        workbook.close()
